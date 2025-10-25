# -*- coding: utf-8 -*-
from enum import Enum
from flask import Flask, render_template, request, flash, redirect, url_for
from markupsafe import Markup
from flask_wtf import FlaskForm, CSRFProtect
from wtforms.validators import DataRequired, Length, Regexp, InputRequired
from wtforms.fields import *
from flask_bootstrap import Bootstrap5, SwitchField
import kanboard
import datetime
import base64
import subprocess
import time 

from pathlib import Path

from odfdo import Document

from openpyxl import Workbook
from openpyxl import load_workbook

import os
from dotenv import load_dotenv

load_dotenv()


kanboard_token = os.getenv('KANBOARD_TOKEN')
kanban_board_name = os.getenv('KANBOARD_BOARD_NAME')
kanban_board_api_point = os.getenv('KANBOARD_ENDPOINT',"http://localhost:8880/jsonrpc.php")
max_repairtime = int(os.getenv('REPARATUR_TIME',40))
repair_guys = int(os.getenv('REPARATEURE',10))
print_active = int(os.getenv('PRINT_AUTO',0))
color_default = os.getenv('COLOR_DEFAULT',"blue")
color_textil = os.getenv('COLOR_TEXTIL',"green")
public_link = os.getenv('READONLY_LINK')

print("Public Link = " + public_link )

min_waiting_time = 5

app = Flask(__name__)
app.secret_key = 'dev'
app.config['WTF_CSRF_TIME_LIMIT']=7200

excel_file = "RepairCafe_Okt_2025.xlsx"
WORKSHEET_TITLE = "RepairCafe Okt 25"
SOURCE = "Reparaturblatt_A4_template.odt"
TARGET = "Reparaturblatt_Okt_2025_Nr"
EXACT_DATE = "18.10.2025"


bootstrap = Bootstrap5(app)
csrf = CSRFProtect(app)


class RepairCafeForm(FlaskForm):
    first_name = StringField("Vorname")
    last_name = StringField("Nachname *", validators=[DataRequired(), Length(2, 100)])
    city = StringField("Wohnort *", validators=[DataRequired(), Length(2, 100)])
    phone = TelField("Telefon")
    email = EmailField("Email")
    age_radio_option = RadioField("Altersgruppe *", validators=[InputRequired()], choices =[("age_0_20","0 - 20 Jahre"),("age_20_40","21 - 40 Jahre"),("age_40_60","41 - 60 Jahre"),("age_60_more","60+ Jahre")])
    turbine_mailinglist = BooleanField('Ich möchte über die nächsten Repair-Cafes in der Turbine informiert werden')
    konsumenten_mailinglist = BooleanField('Der Konsumentenschutz darf mich über Repair-Themen informieren')
    info_newspaper = BooleanField('Zeitung')
    info_poster = BooleanField('Plakataushang')
    info_social_media = BooleanField('Social Media')
    info_website = BooleanField('Newsletter/Website')
    info_mouth = BooleanField('Mund zu Mund')
    info_other = BooleanField('Anderes')
    info_other_what = StringField("")
    repair_object_type = StringField("Gegenstand *", validators=[DataRequired(), Length(2, 200)])
    repair_object_brand = StringField("Hersteller")
    repair_object_category = RadioField("Kategorie *", validators=[InputRequired()], choices=[('handy', 'Handy/Tablet/Unterhaltungselektronik'), ('computer', 'Computer/Laptop'), ('haushalt', 'Haushaltsgeräte'), ('holz_etc', 'Holz/Metall/Mechanik'), ("textil", "Textilien/Leder"), ("sonstiges", "Übriges/Allerlei")])
    repair_object_error = TextAreaField("Fehler-Beschreibung, bitte so genau wie möglich! *", validators=[DataRequired(), Length(2, 1000)])
    submit1 = SubmitField("Diesen Gegenstand registrieren")
    submit2 = SubmitField("Weitere Gegenstände registrieren")


class ConfigForm(FlaskForm):
    max_repairtime = RadioField("Maximale Reparaturzeit", choices = [("30", "30 min"),("40","40 min"),("50","50 min"),("60", "1 Stunde")])
    repair_guys = RadioField("Anzahl Reparierende", choices = [("1","1"),("2","2"),("3","3"),("4","4"),("5","5"),("6","6"),("7","7"),("8","8"),("9","9"),("10","10")])
    print_active = BooleanField('Druck automatisch')
    submit = SubmitField("Konfiguration sichern")

def WriteExcelHeader ():
    try: 
        workbook = load_workbook(filename = excel_file )
    except FileNotFoundError: 
        workbook = Workbook()
        
    worksheet = workbook.active

    worksheet.title=WORKSHEET_TITLE

    worksheet.cell(row=1,column=1,value="Name")
    worksheet.cell(row=1,column=2,value="Vorname")
    worksheet.cell(row=1,column=3,value="Wohnort")
    worksheet.cell(row=1,column=4,value="Telefon")
    worksheet.cell(row=1,column=5,value="Alter")
    worksheet.cell(row=1,column=6,value="Email")
    worksheet.cell(row=1,column=7,value="Turbinen-Mail")
    worksheet.cell(row=1,column=8,value="Konsumenten-Mail")
    worksheet.cell(row=1,column=9,value="Zeitung")
    worksheet.cell(row=1,column=10,value="Plakat")
    worksheet.cell(row=1,column=11,value="Social Media")
    worksheet.cell(row=1,column=12,value="Newsletter")
    worksheet.cell(row=1,column=13,value="Gegenstand")
    worksheet.cell(row=1,column=14,value="Hersteller")
    worksheet.cell(row=1,column=15,value="Kategorie")
    worksheet.cell(row=1,column=16,value="Defekt")
    worksheet.cell(row=1,column=17,value="Repariert?")

    workbook.save(excel_file)

def WriteExcelEntry( number, form ):

    workbook = load_workbook(filename = excel_file )
    worksheet = workbook.active
    worksheet.cell(row=number + 1,column=1,value=form.last_name.data )
    worksheet.cell(row=number + 1,column=2,value=form.first_name.data )
    worksheet.cell(row=number + 1,column=3,value=form.city.data )
    worksheet.cell(row=number + 1,column=4,value=form.phone.data )
    worksheet.cell(row=number + 1,column=5,value=form.age_radio_option.data )  
    worksheet.cell(row=number + 1,column=6,value=form.email.data )
    worksheet.cell(row=number + 1,column=7,value=form.turbine_mailinglist.data )
    worksheet.cell(row=number + 1,column=8,value=form.konsumenten_mailinglist.data )
    worksheet.cell(row=number + 1,column=9,value=form.info_newspaper.data )
    worksheet.cell(row=number + 1,column=10,value=form.info_poster.data )
    worksheet.cell(row=number + 1,column=11,value=form.info_social_media.data )
    worksheet.cell(row=number + 1,column=12,value=form.info_website.data )
    worksheet.cell(row=number + 1,column=13,value=form.repair_object_type.data )
    worksheet.cell(row=number + 1,column=14,value=form.repair_object_brand.data )
    worksheet.cell(row=number + 1,column=15,value=form.repair_object_category.data )
    worksheet.cell(row=number + 1,column=16,value=form.repair_object_error.data )
    workbook.save(excel_file)

def base64_encode_file( file_name ):

    with open(file_name, 'rb') as binary_file:
        binary_file_data = binary_file.read()
        base64_encoded_data = base64.b64encode(binary_file_data)
        base64_output = base64_encoded_data.decode('utf-8')

        return base64_output


def attach_file_to_task ( task_id, file_name ):
    kb = kanboard.Client(kanban_board_api_point, 'jsonrpc',kanboard_token )
    project_props = kb.get_project_by_name(name=kanban_board_name)
    base64_file = base64_encode_file(file_name)
    kb.create_task_file( project_id=project_props["id"], 
                         task_id=task_id,
                         filename=file_name,
                         blob=base64_file)

def create_new_task_on_board(form):
    try:
        kb = kanboard.Client(kanban_board_api_point, 'jsonrpc',kanboard_token )
        project_props = kb.get_project_by_name(name=kanban_board_name)
        color = color_textil  if form.repair_object_category.data == "textil" else color_default
        task_id = kb.create_task(project_id=project_props["id"], 
                                 title=form.repair_object_brand.data + " : " + form.repair_object_type.data, 
                                 description = "# Besitzer  \n\n" + form.last_name.data + "\n# Fehler: \n\n" + form.repair_object_error.data,
                             color_id=color)
        return task_id
    except:
        return 0

def get_amount_waiting_tasks( ):
    kb = kanboard.Client(kanban_board_api_point, 'jsonrpc',kanboard_token )
    project_props = kb.get_project_by_name(name=kanban_board_name)
    column_props = kb.get_columns(project_id=project_props["id"])
    waiting_tasks = kb.search_tasks(project_id=project_props["id"], query="column:" + "\"" + column_props[1]["title"] + "\"" )
    return len(waiting_tasks)

def get_amount_active_tasks( ):
    try:
        kb = kanboard.Client(kanban_board_api_point, 'jsonrpc',kanboard_token )
        project_props = kb.get_project_by_name(name=kanban_board_name)
        column_props = kb.get_columns(project_id=project_props["id"])
        active_tasks = kb.search_tasks(project_id=project_props["id"], query="column:" + "\"" + column_props[2]["title"] + "\"" )
        return len(active_tasks)
    except:
        return -1

def get_active_time():
    kb = kanboard.Client(kanban_board_api_point, 'jsonrpc', kanboard_token )
    project_props = kb.get_project_by_name(name=kanban_board_name)
    column_props = kb.get_columns(project_id=project_props["id"])
    active_tasks = kb.search_tasks(project_id=project_props["id"], query="column:" + "\"" + column_props[2]["title"] + "\"" )
    now = datetime.datetime.now()
    summ_in_minutes = 0 
    for active_task in active_tasks:
        starttime = datetime.datetime.fromtimestamp(active_task["date_moved"]) 
        duration = now - starttime
        duration_minutes = int( duration.total_seconds() / 60 )
        if duration_minutes  > max_repairtime:
            duration_minutes = max_repairtime
        summ_in_minutes = summ_in_minutes + duration_minutes 
    return int(summ_in_minutes)

def get_waiting_time( ):
    waiting_tasks = get_amount_waiting_tasks()
    active_tasks = get_amount_active_tasks()
    active_tasks_active_time = get_active_time()

    if waiting_tasks + active_tasks < repair_guys :
        return min_waiting_time
    else:
        overall_waiting_time = ( waiting_tasks + active_tasks ) *  max_repairtime 
        overall_waiting_time = overall_waiting_time - active_tasks_active_time
        average_waiting_time = overall_waiting_time / int (repair_guys)
        average_waiting_time = int( int( int(average_waiting_time / 5 ) + 1) * 5)
        if average_waiting_time < min_waiting_time:
            average_waiting_time = min_waiting_time
        return average_waiting_time 
            




def save_new(document: Document, name: str):
    new_path = name
    print("Saving:", new_path)
    document.save(new_path, pretty=True)

def create_new_document( form, number ):

    document = Document(SOURCE)
    target_name = TARGET + "_" + str(number) + ".odt"
    body = document.body

    body.replace("11.11.1111", EXACT_DATE)
    body.replace("2222", str(number) )
    body.replace("Hugo Egon", form.first_name.data + " " + form.last_name.data )
    body.replace("Lichtschwert", form.repair_object_type.data )
    body.replace("Force", form.repair_object_brand.data )
    body.replace("Kristall", form.repair_object_error.data)
    body.replace("hugo@egon", form.email.data)
    if form.info_newspaper.data : body.replace("□ Zeitung", "☑ Zeitung") 
    if form.info_poster.data  : body.replace("□ Plakataushang", "☑ Plakataushang")     
    if form.info_social_media.data  : body.replace("□ Internet / Soziale Medien", "☑ Internet / Soziale Medien")     
    if form.info_website.data  : body.replace("□ Newsletter", "☑ Newsletter/Website")     
    if form.turbine_mailinglist.data  : body.replace("□ Repaircafes in der Turbine", "☑ Repaircafes in der Turbine")     
    if form.konsumenten_mailinglist.data : body.replace("□ Reparaturthemen allgemein", "☑ Reparaturthemen allgemein")     
    save_new(document, target_name )
    return target_name

def print_document( filename ):
    if print_active == 1 : 
        subprocess.Popen(["libreoffice", "--norestore", "-p", filename ])

@app.route('/oldindex')
def index():
    return render_template('index.html')


@app.route('/board', methods=['GET', 'POST'])
def board():
    return render_template('board.html')

@app.route('/publicboard', methods=['GET', 'POST'])
def publicboard():
    global public_link
    return render_template('publicboard.html', url=public_link)

@app.route('/overview', methods=['GET', 'POST'])
def overview():
    active = get_amount_active_tasks()
    if active == -1:
        flash('ACHTUNG! Bitte Administrator benachrichtigen, keine Verbindung zum Kanban-Board')
        active = "n/a"
        waiting = "n/a"
        waiting_time = "n/a"
    else:    
        waiting = get_amount_waiting_tasks()
        waiting_time = get_waiting_time ()  
    return render_template('overview.html', active=str(active), queued=str(waiting),waiting_time = str(waiting_time), repair_guys = str(repair_guys),max_repairtime = str(max_repairtime)   )

@app.route('/toggle')
def toggle():
    """
    Alternates between returning the response from /publicboard and /overview 
    every 10 seconds
    """
    # Get the current time in seconds since the epoch
    current_time_seconds = time.time()


    # Determine which 10-second block we are in.
    # Integer division by 10 gives us the block number.
    
    ten_seconds_block_index = current_time_seconds // 10

    # Check if the block index is even or odd.
    # 0, 2, 4... serve route 1
    # 1, 3, 5... serve route 2
    if ten_seconds_block_index % 2 == 0:
        # Even block: call the function handling route1 directly
        # print(f"Time {time.ctime(current_time_seconds)}: Serving Route 1 (Block {ten_seconds_block_index})")
        # We call the view function directly to get its return value
        return overview()
    else:
        # Odd block: call the function handling route2 directly
        # print(f"Time {time.ctime(current_time_seconds)}: Serving Route 2 (Block {ten_seconds_block_index})")
        # We call the view function directly to get its return value
        return publicboard()


@app.route('/', methods=['GET', 'POST'])
def register_form():
    form = RepairCafeForm()
    if form.validate_on_submit():
        rep_nr = create_new_task_on_board(form)
        if rep_nr == 0:
            flash('ACHTUNG! Gegenstand konnte nicht registriert werden! Bitte Administrator benachrichtigen, keine Verbindung zum Kanban-Board')
        else:
            flash('Reparatur für : ' + form.repair_object_type.data + " wurde mit der Nummer : " + str(rep_nr) + "  registriert! Bitte mit dieser Nummer zur Registration.   Bitte Ausgabe schliessen ==> " )
            filename = create_new_document( form, rep_nr )
            attach_file_to_task( rep_nr, filename ) 
            WriteExcelEntry( rep_nr , form )
            print_document( filename )
        return redirect(url_for('register_form'))
    return render_template(
        'form.html',
        form=form,
        repaircafe_form=RepairCafeForm()
    )

@app.route('/config', methods=['GET', 'POST'])
def config_form():
    global max_repairtime
    global repair_guys
    global print_active
    print("Before: Repairtime = " + str(max_repairtime) + "Guys = " + str(repair_guys) )
    form = ConfigForm()
    if form.validate_on_submit():
        flash('Konfiguration registriert !')
        max_repairtime = int(form.max_repairtime.data)
        repair_guys = int(form.repair_guys.data)
        print_active = int(form.print_active.data)
        print("After: Repairtime = " + str(max_repairtime) + "Guys = " + str(repair_guys) )
        return redirect(url_for('overview'))
    return render_template(
        'config.html',
        form=form,
        config_form=ConfigForm()
    )
if __name__ == '__main__':
    
    WriteExcelHeader ()
    app.run(debug=True, host='0.0.0.0', port=80 )

